import math

from openpyxl.workbook import Workbook
from pandas import *
import matplotlib.pyplot as plt
from math import radians, cos, sin, asin, sqrt
import numpy as np
import scipy

parameter_filter = 40
output = Workbook()
ws = output.active
language_lookup = {}
language_dictionary = {}
language_names = read_csv("languages.csv")
values = read_csv('values.csv')
names = language_names['Name'].tolist()
codes = language_names['ID'].tolist()
latitudes = language_names['Latitude'].tolist()
longitudes = language_names['Longitude'].tolist()
language_id = values['Language_ID'].tolist()
parameters = values['Parameter_ID'].tolist()
parameter_values = values['Value'].tolist()
families = language_names['Family'].tolist()

for index in range(0, len(codes)):
    language_lookup[codes[index]] = names[index]

for language in names:
    language_dictionary[language] = {'available_parameters': [], 'given_values': [], 'latitude': latitudes[names.index(language)], 'longitude': longitudes[names.index(language)], 'family': families[names.index(language)]}

for value_row in values.iterrows():
    value_name = language_lookup[value_row[1][1]]
    language_dictionary[value_name]['available_parameters'].append(value_row[1][2])
    language_dictionary[value_name]['given_values'].append(value_row[1][3])

maximums = read_csv('codes.csv')
code_max = maximums['Parameter'].tolist()
max_values = maximums['Max'].tolist()
def value_comparison(target, compare):
    percentages = []
    avail_index = 0
    for avail_code in language_dictionary[target]['available_parameters']:
        if avail_code in language_dictionary[compare]['available_parameters']:
            max_index = code_max.index(avail_code)
            maximum = max_values[max_index]
            compare_index = language_dictionary[compare]['available_parameters'].index(avail_code)
            avail_value = language_dictionary[target]['given_values'][avail_index]
            compare_value = language_dictionary[compare]['given_values'][compare_index]
            percentages.append((1-(abs(compare_value-avail_value))/(maximum-1))**2)

        avail_index += 1
    if(len(percentages) > parameter_filter):
        return(sum(percentages)/len(percentages))
    else:
        return 0


def haversine(lat1, lon1, lat2, lon2):
    lon1 = radians(lon1)
    lat1 = radians(lat1)
    lon2 = radians(lon2)
    lat2 = radians(lat2)

    dlon = lon2 - lon1
    dlat = lat2 - lat1
    a = sin(dlat/2)**2 + cos(lat1) * cos(lat2) * sin(dlon/2)**2
    distance = 2 * asin(sqrt(a)) * 3958.8

    return distance


def comparison_function(target_language):
    comparisons = []
    comparison_languages = []
    distances = []
    for compare_language in names:
        percentage = value_comparison(target_language, compare_language)
        if percentage != 0 and percentage != 1.0:
            comparison_languages.append(compare_language)
            comparisons.append(percentage)
            distances.append(haversine(language_dictionary[target_language]['latitude'], language_dictionary[target_language]['longitude'], language_dictionary[compare_language]['latitude'], language_dictionary[compare_language]['longitude']))
    if comparisons == []:
        comparisons = [0]
        comparison_languages = [0]
        distances = [0]
    return comparisons, comparison_languages, distances

def regression(comparisons, distances):
    x = np.array(distances)
    y = np.array(comparisons)
    slope, intercept, r_value, p_value, std_err = scipy.stats.linregress(x, np.log(y))
    line = np.linspace(0, 12000, 5)
    print(slope, intercept, r_value ** 2)
    plt.scatter(x, y)
    plt.plot(line, math.e**(slope * line + intercept))
    plt.show()

mechanism = input("Manual or Automatic?")


if mechanism == "Manual":
    target_language = input("Please input a language")
    comparisons, comparison_languages, distances = comparison_function(target_language)
    print(
        "Min: " + str(min(comparisons)) + ", " + comparison_languages[comparisons.index(min(comparisons))] + ", " + str(
            distances[comparisons.index(min(comparisons))]))
    print(
        "Max: " + str(max(comparisons)) + ", " + comparison_languages[comparisons.index(max(comparisons))] + ", " + str(
            distances[comparisons.index(max(comparisons))]))
    regression(comparisons, distances)

elif mechanism == "Automatic":
    language_list = []
    family_list = []
    min_distance_list = []
    max_distance_list = []
    min_names_list = []
    max_names_list = []
    min_percentage_list = []
    max_percentage_list = []
    for language in names:
        if len(language_dictionary[language]['available_parameters']) > parameter_filter:
            print(language)
            comparisons, comparison_languages, distances = comparison_function(language)
            min_distance_list.append(distances[comparisons.index(min(comparisons))])
            max_distance_list.append(distances[comparisons.index(max(comparisons))])
            min_percentage_list.append(min(comparisons))
            max_percentage_list.append(max(comparisons))
            min_names_list.append(comparison_languages[comparisons.index(min(comparisons))])
            max_names_list.append(comparison_languages[comparisons.index(max(comparisons))])
            language_list.append(language)
            family_list.append(language_dictionary[language]['family'])
    for spreadsheet_index in range(0, len(language_list)):
        ws['A' + str(spreadsheet_index + 1)] = language_list[spreadsheet_index]
        ws['B' + str(spreadsheet_index + 1)] = min_percentage_list[spreadsheet_index]
        ws['C' + str(spreadsheet_index + 1)] = min_distance_list[spreadsheet_index]
        ws['D' + str(spreadsheet_index + 1)] = min_names_list[spreadsheet_index]
        ws['E' + str(spreadsheet_index + 1)] = max_percentage_list[spreadsheet_index]
        ws['F' + str(spreadsheet_index + 1)] = max_distance_list[spreadsheet_index]
        ws['G' + str(spreadsheet_index + 1)] = max_names_list[spreadsheet_index]
        ws['H' + str(spreadsheet_index + 1)] = family_list[spreadsheet_index]
    output.save("output.xlsx")





